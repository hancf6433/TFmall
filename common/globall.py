# -*- encoding=utf-8 -*-
'''
全局对象
'''
import os
import re
import time
from typing import Union
import openpyxl
from airtest.core.android.adb import ADB
from poco.drivers.ios import iosPoco
from poot.core.api import Poot
import common.setting as ST
from poco.pocofw import Poco
from airtest.utils.logger import get_logger
from airtest.core.api import (init_device, connect_device, set_current, auto_setup)
from poco.drivers.android.uiautomation import AndroidUiautomationPoco
from common.tools import tools
import threading

LOGGING = get_logger("airtest.||%s||" % __name__)
endFlag=False
class TestData(object):

    def __init__(self):
        '''
        测试数据专用类
        '''
        self.__data = {}  # type:dict #测试数据
        self.__groupList = set()  # type:set #数据分组列表
        self.__now_group = None  # type:str #当前使用分组


    @property
    def group(self):
        '''
        返回当前使用分组
        :return:
        '''
        return self.__now_group

    @group.setter
    def group(self, groupName: str):
        '''
        设置当前使用分组
        :param groupName:
        :return:
        '''
        if groupName not in self.__groupList:
            raise BaseException("不存在此分组{%s}" % groupName)
        self.__now_group = groupName

    @property
    def groupList(self):
        '''
        返回所有分组
        :return:
        '''
        return self.__groupList

    @groupList.setter
    def groupList(self, groupName):
        '''
        增加数据分组
        :param groupName:
        :return:
        '''
        self.groupList.add(groupName)

    def put_test_data(self, key, value, groupName=None):
        '''
        放置测试数据
        :param key:
        :param value:
        :param groupName:
        :return:
        '''
        if groupName == None:
            groupName = self.group
        if self.__data.get(groupName):
            # 存在数据分组
            tempData: dict = self.__data.get(groupName)
            tempData[key] = value
        else:
            # 不存在数据分组
            # 增加数据分组
            self.__groupList.add(groupName)
            tempData: dict = {}
            tempData[key] = value
            self.__data[groupName] = tempData

    def get_test_data(self, key, groupName=None):
        '''
        获取测试数据
        :param key:
        :param groupName:
        :return:
        '''
        if groupName == None:
            groupName = self.group
        if self.__data.get(groupName):
            tempData: dict = self.__data.get(groupName)
            if not tempData.get(key):
                raise BaseException("{%s}数据项不存在于{%s}数据分组" % (key, groupName))
            return tempData.get(key)
        else:
            raise BaseException("不存在数据分组：{%s}" % groupName)



class g(object):
    '''
    全局对象,主要用来进行设备的初始化和加载，及相关全局信息存储。
    '''

    def __new__(cls, *args, **kwargs):
        cls._flag = False
        if not hasattr(cls, "_instance"):
            cls._flag = True
            cls._instance = super().__new__(cls, *args, **kwargs)
        return cls._instance

    def __init__(self):
        if self._flag:
            #当前活动的设备名字
            self.__activeDevName=""
            #电器活动的设备，一个airtest对象
            self.__activeDev = None
            # 当前活动的设备，一个Poco对象
            self.__activePoco = None  # type:Poco
            # 当前活动设备类型
            self.__activeDevType = ""  # type:str
            # 脚本开始时间
            self.__beginTime = ""  # type:str
            # 脚本结束时间
            self.__endTime = ""  # type:str
            # 脚本持续时间
            self.__duration=""
            # airtest对象实例
            self.__airtestDevDict = dict()
            # poco对象列表
            self.__pocoDevList = dict()
            # 设备名字和设备实例映射
            self.__nameAndDevMap = dict()
            # 所有case运行状态字典
            self.__data=dict()
            # 单个报告信息列表
            self.__report_list=list()
            # 报告存储路径
            self.__real_log_path=""
            # 测试用数据
            self.__test_data=TestData()
            # 资源信息
            self.__resource_infor=dict()
            # 图片信息
            self.__picture=dict()
            self.__android="Android"
            self.__ios="IOS"
            # 有名字的变量正则
            self.__var_name_pattern=re.compile(r'\$\{([0-9a-zA-Z]+)\}')
            self.__var_pattern=re.compile(r'\$\{\}')
            # 未killcase执行次数
            self.__run_count=0
            # 当前登录用户手机号
            self.__now_phone_number=None
            # 储存模块
            self.__save_module=None
            #poot引用
            self.__activePoot=None
            #poot存储
            self.__pootDevList=dict()
            #使用poot
            self.__userPoot=False
            # 实际模块
            self.__module=list()
            self.loadTestData()
            self.initDevice()


    def isThePootUsed(self):
        '''
        返回poot
        :return:
        '''
        return self.__userPoot


    def getDictByKeyName(self,dicName,keyName):
        '''
        返回key，value
        :param dicName: 字典名
        :param keyName: key
        :return:
        '''
        pass


    def loadTestData(self):
        '''
        加载数据文件
        :return:
        '''
        if ST.ANDROID_DIR:
            #加载安卓资源信息
            #读取excel的方法(弃用）
            # androidSheets=self.__getFileSheetName(ST.ANDROID_DIR)
            # for sheetName in androidSheets:
            #     self.__save_module=sheetName.lower()
            #     self.__loadDataFromSheet(ST.ANDROID_DIR,sheetName,self.__put_android_resource_infor)
            #读取xml的方法
            for root, paths, files in os.walk(ST.ANDROID_DIR):
                for file in files:
                    module = os.path.splitext(file)[0].lower()
                    realFile = os.path.join(root, file)
                    fileModule,uiDict=tools.resolveUiXml(realFile)
                    if module!=fileModule:
                        raise BaseException("%s的文件名与根节点name属性值不一致：%s" % (file,fileModule))
                    self.__save_module=fileModule
                    for key,value in uiDict.items():
                        self.__put_android_resource_infor(key,value)
        if ST.IOS_RESOURCE_DIR:
            #加载ios资源信息
            # iosSheets=self.__getFileSheetName(ST.IOS_RESOURCE_DIR)
            # for sheetName in iosSheets:
            #     self.__save_module=sheetName.lower()
            #     self.__loadDataFromSheet(ST.IOS_RESOURCE_DIR,sheetName,self.__put_ios_resource_infor)
            # 读取xml的方法
            for root, paths, files in os.walk(ST.IOS_RESOURCE_DIR):
                for file in files:
                    module = os.path.splitext(file)[0].lower()
                    realFile = os.path.join(root, file)
                    fileModule, uiDict = tools.resolveUiXml(realFile)
                    if module != fileModule:
                        raise BaseException("%s的文件名和根节点名不一致：%s" % (file, fileModule))
                    self.__save_module = fileModule
                    for key, value in uiDict.items():
                        self.__put_ios_resource_infor(key, value)
        if ST.PICTURE_DIR:
            #加载图片信息
            for root, paths, files in os.walk(ST.PICTURE_DIR):
                for file in files:
                    file = os.path.join(root, file)
                    module = os.path.basename(os.path.dirname(file)).lower()
                    key = os.path.basename(file)
                    value = file
                    self.__save_module=module
                    self.__put_picture(key,value)
        #加载测试数据
        if ST.TEST_DATA_FILE :
            # testSheets=self.__getFileSheetName(ST.TEST_DATA_FILE)
            # for sheetName in testSheets:
            #     self.__test_data.groupList=sheetName
            #     self.__test_data.group=sheetName
            #     self.__loadDataFromSheet(ST.TEST_DATA_FILE,sheetName,self.__test_data.put_test_data)
            # if ST.TEST_DATA_USE_SHEET_NAME:
            #     self.__test_data.group=ST.TEST_DATA_USE_SHEET_NAME
            dataDic=tools.resolveTestDataXml(ST.TEST_DATA_FILE)
            for dataSetName,datas in dataDic.items():
                self.__test_data.groupList=dataSetName
                self.__test_data.group=dataSetName
                for key,value in datas.items():
                    self.__test_data.put_test_data(key,value)

    def __getFileSheetName(self,file):
        '''
        获取一个文件的所有sheet名
        :return:
        '''
        excel=None
        try:
            excel=openpyxl.load_workbook(file,read_only=True)
            sheetNames=[]
            for sheetName in excel.sheetnames:
                sheetNames.append(sheetName)
            return sheetNames
        except FileNotFoundError:
            raise BaseException("文件不存在，请检查：{%s}" % file)
        except PermissionError:
            raise BaseException("文件加载失败，请检查是否被占用或打开，文件名：{%s}" % file)
        except:
            raise BaseException("文件加载失败，文件名：{%s}" % file)
        finally:
            if excel:
                excel.close()


    def __loadDataFromSheet(self,filePath,sheetName,putDataMethod):
        '''
        载入单个sheet的数据
        :param filePath: 数据文件路径
        :param sheetName: sheet名
        :param putDataMethod: 执行存储的方法
        :return:
        '''
        data_file =filePath
        if data_file == None:
            return
        excel=None
        try:
            excel=openpyxl.load_workbook(data_file, read_only=True)
            sheet = excel.get_sheet_by_name(sheetName)
            for row in sheet.rows:
                if row[0].value == ST.SHEET_KEY_NAME:
                    continue
                key = str(row[0].value).strip()
                value = str(row[1].value).strip()
                if key == "None" or key == "":
                    continue
                if value == "None" or value == "":
                    continue
                putDataMethod(key, value)
        except KeyError:
            raise BaseException("{%s}sheet不存在于{%s}" % (sheetName,data_file))
        except PermissionError:
            raise BaseException("{%s}数据文件加载失败,请检查文件是否被占用或是被打开。" % (data_file,))
        except:
            raise BaseException("{%s}数据文件加载失败,Sheet：{%s}" % (data_file,sheetName))
        finally:
            if excel:
                excel.close()

    def initDevice(self):
        '''
        初始化设备
        :return:
        '''

        if (ST.DEVICE is None) or (len(ST.DEVICE) == 0):
            #未配置设备时，使用默认设备进行连接
            #先尝试连接安卓设备，如果连接失败，连接IOS
            try:
                dev = init_device(cap_method=ST.CAP_METHOD, ori_method=ST.ORI_METHOD, touch_method=ST.TOUCH_METHOD)
            except:
                dev=init_device(self.__ios)
            self.__airtestDevDict["___one___"]=dev
            self.__nameAndDevMap["___one___"]=dev.uuid
            devType=dev.__class__.__name__
            if devType==self.__android:
                pocoDev = AndroidUiautomationPoco()
                pootDev=Poot(dev.uuid,screenshot_each_action=True)
            elif devType==self.__ios:
                pocoDev=iosPoco()
                pootDev=None
            else:
                raise BaseException("暂时不支持此种类型设备{%s}" % devType)
            self.__pocoDevList["___one___"]=pocoDev
            self.__pootDevList["___one___"]=pootDev
            self.__activeDevName="___one___"
            self.__activeDev=dev
            self.__activePoco = pocoDev
            self.__activePoot=pootDev
            self.__activeDevType=devType
        else:
            #否则，使用配置进行连接
            for devName,devUri in ST.DEVICE.items():
                dev=connect_device(devUri)
                self.__airtestDevDict[devName]=dev
                self.__nameAndDevMap[devName]=dev.uuid
                devType = dev.__class__.__name__
                if devType == self.__android:
                    pocoDev = AndroidUiautomationPoco(dev)
                    pootDev=Poot(dev.uuid,screenshot_each_action=True)
                elif devType == self.__ios:
                    pocoDev = iosPoco()
                    pootDev=None
                else:
                    raise BaseException("暂时不支持此种类型设备{%s}" % devType)
                self.__activeDevName = devName
                self.__pocoDevList[devName]=pocoDev
                self.__pootDevList[devName]=pootDev
                self.__activeDev = dev
                self.__activePoco=pocoDev
                self.__activeDevType=devType
                self.__activePoot=pootDev
        LOGGING.info("(%s)已连接" % ",".join(self.__nameAndDevMap.values()))



    def setActiveDev(self,dev:Union[str,int]):
        '''
        根据设备名获得一个初始化的设备poco对象，并将当前airtest操作的设备置为返回的dev
        传入None将获得第一个被初始化的设备
        :param dev:设备名 or 索引
        '''
        connectDevCount=len(self.__airtestDevDict.keys())
        if connectDevCount <= 0:
            raise BaseException("未连接任何设备，请重新连接设备")
        if type(dev)==type("str"):
            # 已配置名字和设备映射
            if self.__nameAndDevMap.get(dev) is None:
                raise BaseException("传入的名称未对应任何设备，请检查是否和配置文件一致")
        elif type(dev)==type(1):
            #传入索引
            if dev >= connectDevCount:
                raise BaseException("传入{%s}索引超出，最大索引为{%s}" %(dev,connectDevCount))
            keys=list(self.__nameAndDevMap.keys())
            dev=keys[dev]
        else:
            raise BaseException("传入参数类型错误，期望str和int，实际{%s}" % type(dev))
        set_current(self.__nameAndDevMap.get(dev))
        # return self.__pocoDevList[devIndex]
        self.__activePoco = self.__pocoDevList.get(dev)
        self.__activePoot=self.__pootDevList.get(dev)
        self.__activeDevType=self.__airtestDevDict.get(dev).__class__.__name__
        self.__activeDev=self.__airtestDevDict.get(dev)
        self.__activeDevName=dev

    @property
    def activePoco(self)->Poco:
        '''
        默认激活的设备是配置文件里的最后一个设备，如果没有配置文件，将直接选取连接的设备
        :return:
        '''
        if self.__userPoot:
            self.__userPoot=False
            pocoDev=AndroidUiautomationPoco(self.__activeDev)
            self.__pocoDevList[self.__activeDevName]=pocoDev
            self.__activePoco=pocoDev
        return self.__activePoco

    @property
    def activePoot(self)->Poot:
        '''
         默认激活的设备是配置文件里的最后一个设备，如果没有配置文件，将直接选取连接的设备
        :return:
        '''
        if not self.__userPoot:
            self.__userPoot=True
            self.__activePoco.stop_running()
        return self.__activePoot

    @property
    def activeDev(self):
        '''
        默认激活
        :return:
        '''
        return self.__activeDev

    @property
    def data(self)->dict:
        '''
        返回脚本数据存储列表
        :return:
        '''
        return self.__data

    @property
    def beginTime(self):
        '''
        设置脚本开始时间
        :return:
        '''
        return self.__beginTime

    @beginTime.setter
    def beginTime(self,dateTime:str):
        '''
        返回脚本开始时间
        :param dateTime:
        :return:
        '''
        self.__beginTime=dateTime

    @property
    def endTime(self):
        '''
        设置脚本开始时间
        :return:
        '''
        return self.__endTime


    @endTime.setter
    def endTime(self,dateTime:str):
        '''
        返回脚本开始时间
        :param dateTime:
        :return:
        '''
        self.__endTime=dateTime

    @property
    def duration(self):
        '''
        返回持续时间
        :return:
        '''
        return self.__duration

    @duration.setter
    def duration(self,value):
        '''
        设置持续时间
        :param value:
        :return:
        '''
        self.__duration=value

    @property
    def report_list(self):
        '''
        返回报告信息列表
        :return:
        '''
        return self.__report_list

    @property
    def real_log_path(self):
        '''
        返回报告路径
        :return:
        '''
        return self.__real_log_path

    @real_log_path.setter
    def real_log_path(self,path:str):
        '''
        设置报告路径
        :param path:
        :return:
        '''
        self.__real_log_path=path

    @property
    def group(self):
        '''
        返回当前数据分组
        :return:
        '''
        return self.__test_data.group

    @group.setter
    def group(self,groupName):
        '''
        设置数据分组
        :param groupName:
        :return:
        '''
        self.__test_data.group=groupName

    def get_test_data(self,key,groupName=None):
        '''
        获取测试数据
        :param key:
        :return:
        '''
        return self.__test_data.get_test_data(key,groupName)



    def put_test_data(self,key,value,groupName=None):
        '''
        放置测试数据
        :param key:
        :param value:
        :return:
        '''
        self.__test_data.put_test_data(key,value,groupName)

    def __getInforFromDict(self,inforDict:dict,keyList:list):
        '''
        从信息字典中嵌套取出信息，如果某层信息不存在则返回None
        :param inforDict:
        :param keyList:
        :return:
        '''
        for key in keyList:
            if not inforDict:
                return None
            if type(inforDict)!=dict:
                raise BaseException("必须是一个字典")
            inforDict=inforDict.get(key)
        return inforDict



    def get_resource_infor(self,key,*args,**kw):
        '''
        获取资源信息
        :param key:
        :param args:顺序参数   ${}
        :param kw:有name名的参数    ${name}
        :return:
        '''
        type=self.__activeDevType
        value=None
        tempM=None
        #取出模块
        for module in self.__module:
            tempM=self.__getInforFromDict(self.__resource_infor,[type,module.lower()])
            if tempM:
                break
        if not tempM:
            raise BaseException("在{%s}的UI信息未找到{%s}模块，请检查" % (type,self.__module))
        #取出value值
        for module in self.__module:
            value=self.__getInforFromDict(self.__resource_infor,[type,module.lower(),key])
            if value:
                break
        if not value:
            raise BaseException("在{%s}的UI信息的{%s}模块里未找到UI名称为{%s}的定位信息，请检查文件。" % (type,self.__module,key))
        #先进行变量检查
        orderMatcher=self.__var_pattern.findall(value)
        nameMatcher=self.__var_name_pattern.findall(value)
        if len(orderMatcher)!=len(args):
            raise BaseException("需要的顺序变量数{%s},实际有{%s}" % (len(orderMatcher), len(args)))
        if len(nameMatcher) != len(kw):
            raise BaseException("需要的命名变量数{%s},实际有{%s}" % (len(nameMatcher), len(kw)))
        #先替换顺序变量
        if args!=None:
            for a in args:
                value = value.replace("${}", a, 1)
        #再替换命名变量
        if kw!=None:
            for k,v in kw.items():
                # 提取value中的变量
                # #如果变量名存在于字符串中，则进行替换
                if k in nameMatcher:
                    value=value.replace("${%s}" % k,str(v))
        return value

    def __put_android_resource_infor(self,key,value):
        '''
        放置安卓资源信息
        :param key:
        :param value:
        :return:
        '''
        self.__put_resource_infor(key,value,self.__android)

    def __put_ios_resource_infor(self,key,value):
        '''
        放置ios资源信息
        :param key:
        :param value:
        :return:
        '''
        self.__put_resource_infor(key,value,self.__ios)

    def __put_resource_infor(self,key,value,type):
        '''
        放置资源信息
        :param key:
        :param value:
        :param type: IOS or ANDROID
        :return:
        '''
        temp=self.__resource_infor.get(type)
        if not temp:
            temp={}
        module=temp.get(self.__save_module)
        if not module:
            module={}
        module[key]=value
        temp[self.__save_module]=module
        self.__resource_infor[type]=temp

    def get_picture(self,key):
        '''
        获取图片识别信息
        :param key:
        :return:
        '''
        value=None
        if ".png" not in key:
            key+=".png"
        for module in self.__module:
            value=self.__getInforFromDict(self.__picture,[module.lower(),key])
        if not value:
            raise BaseException("图片识别目录{%s}不存在图片数据：{%s}" % (self.__module,key))
        return value



    def __put_picture(self,key,value):
        '''
        放置资源信息
        :param key:
        :param value:
        :return:
        '''
        module=self.__picture.get(self.__save_module)
        if not module:
            module={}
        module[key]=value
        self.__picture[self.__save_module]=module

    @property
    def runCount(self):
        self.__run_count+=1
        if self.__run_count>=1:
            self.__run_count=0
        return self.__run_count

    @property
    def userPhone(self):
        '''
        返回当前用户手机号
        :return:
        '''
        return self.__now_phone_number

    @userPhone.setter
    def userPhone(self,phoneNumber):
        '''
        设置当前用户手机号
        :param phoneNumber:
        :return:
        '''
        self.__now_phone_number=phoneNumber

    @property
    def activeDevType(self):
        '''
        返回当前设备的类型
        :return:
        '''
        return self.__activeDevType

    def clearModule(self):
        '''
        清空模块
        :return:
        '''
        self.__module=[]

    def addModule(self,module):
        '''
        新增模块
        :return:
        '''
        self.__module.append(module)

    @property
    def module(self):
        '''
        返回模块列表
        :return:
        '''
        return self.__module

    @module.setter
    def module(self,value:list):
        '''
        设置模块列表
        :param value:
        :return:
        '''
        if type(value)!=list:
            raise BaseException("参数只能是列表类型，不能是{%s}" % type(value))
        self.__module=value
